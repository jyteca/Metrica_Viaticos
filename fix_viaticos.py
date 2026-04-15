"""
Script para crear una versión XLSX corregida del formulario de viáticos.
Recrea todas las hojas con datos + fórmulas funcionales usando openpyxl.
"""

import xlrd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from copy import copy as style_copy

INPUT_FILE = r'Solicitud de Viaticos y Fondos.xls'
OUTPUT_FILE = r'Solicitud de Viaticos y Fondos_CORREGIDO.xlsx'

# Leer datos del archivo original
rb = xlrd.open_workbook(INPUT_FILE)

# Crear nuevo workbook
wb = openpyxl.Workbook()

# Estilos comunes
header_font = Font(name='Arial', size=14, bold=True)
label_font = Font(name='Arial', size=10, bold=True)
data_font = Font(name='Arial', size=10)
money_format = '#,##0'
pct_format = '0%'
date_format = 'DD/MM/YYYY'

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(name='Arial', size=11, bold=True, color='FFFFFF')
total_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

def apply_border(ws, row, col):
    ws.cell(row=row, column=col).border = thin_border

def apply_style(ws, row, col, value=None, font=None, fill=None, fmt=None, alignment=None):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if alignment:
        cell.alignment = alignment
    cell.border = thin_border
    return cell


# =========================================================================
# HOJA 1: Solicitud
# =========================================================================
ws = wb.active
ws.title = 'Solicitud'

# Anchos de columna
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 20

# Titulo
ws.merge_cells('D1:G1')
apply_style(ws, 1, 4, 'FORMULARIO DE SOLICITUD', header_font, header_fill, alignment=Alignment(horizontal='center'))
for c in range(4, 8):
    apply_style(ws, 1, c, font=header_font_white, fill=header_fill)
ws.cell(row=1, column=4).value = 'FORMULARIO DE SOLICITUD'

ws.merge_cells('D2:G2')
apply_style(ws, 2, 4, 'VIATICOS, REMBOLSOS Y GASTOS DE', Font(name='Arial', size=11, bold=True, color='FFFFFF'), header_fill, alignment=Alignment(horizontal='center'))
for c in range(4, 8):
    apply_style(ws, 2, c, fill=header_fill)

ws.merge_cells('D3:G3')
apply_style(ws, 3, 4, 'REPRESENTACION', Font(name='Arial', size=11, bold=True, color='FFFFFF'), header_fill, alignment=Alignment(horizontal='center'))
for c in range(4, 8):
    apply_style(ws, 3, c, fill=header_fill)

# Datos del encabezado - leer del original
rs = rb.sheet_by_index(0)

# Leer campos del original
fields = [
    (5, 'Fecha Solicitud', None),
    (6, 'Cliente', None),
    (7, 'Motivo', None),
    (8, 'ACN', None),
    (9, 'Orden de Compra', None),
    (10, 'Solicitado por', None),
    (11, 'Responsable', None),
]

row = 5
for orig_row, label, _ in fields:
    apply_style(ws, row, 2, label, label_font, light_fill)
    val = rs.cell_value(orig_row - 1, 2)
    if rs.cell_type(orig_row - 1, 2) == 3:  # date
        import datetime
        val = datetime.datetime(*xlrd.xldate_as_tuple(val, rb.datemode))
        apply_style(ws, row, 3, val, data_font, fmt=date_format)
    elif val:
        apply_style(ws, row, 3, str(val), data_font)
    else:
        apply_style(ws, row, 3, '', data_font)
    row += 1

# Tabla de gastos - encabezados
row = 12
headers_gastos = ['Gastos para', 'Viatico no rendible $', 'Fondo a rendir $',
                  'Gastos representacion $', 'Reembolso $', 'Total gastos estimados']
for i, h in enumerate(headers_gastos):
    apply_style(ws, row, 2 + i, h, header_font_white, header_fill, alignment=Alignment(horizontal='center', wrap_text=True))

# Sub-encabezados
row = 13
sub_headers = ['', '$', 'Rendir $', 'Representacion $', '$', 'Estimado']
for i, h in enumerate(sub_headers):
    apply_style(ws, row, 2 + i, h, Font(name='Arial', size=9, italic=True), light_fill, alignment=Alignment(horizontal='center'))

# Filas de gastos con FORMULAS
row = 14
# Peajes (fila 14)
apply_style(ws, 14, 2, 'Peajes', label_font)
apply_style(ws, 14, 3, '', data_font)  # viatico no rendible
apply_style(ws, 14, 4, font=data_font, fmt=money_format)
ws.cell(row=14, column=4).value = "=PEAJE!A2"  # FORMULA -> Peaje total
apply_style(ws, 14, 5, '', data_font)
apply_style(ws, 14, 6, '', data_font)
apply_style(ws, 14, 7, font=data_font, fmt=money_format)

# Envio de carga (fila 15)
apply_style(ws, 15, 2, 'Envio de carga', label_font)
apply_style(ws, 15, 3, '', data_font)
apply_style(ws, 15, 4, 0, data_font, fmt=money_format)  # valor manual
apply_style(ws, 15, 5, '', data_font)
apply_style(ws, 15, 6, '', data_font)
apply_style(ws, 15, 7, font=data_font, fmt=money_format)

# Hotel (fila 16)
apply_style(ws, 16, 2, 'Hotel', label_font)
apply_style(ws, 16, 3, '', data_font)
apply_style(ws, 16, 4, font=data_font, fmt=money_format)
ws.cell(row=16, column=4).value = "='ALOJAMIENTO EN CAMINO'!A2"  # FORMULA
apply_style(ws, 16, 5, '', data_font)
apply_style(ws, 16, 6, '', data_font)
apply_style(ws, 16, 7, font=data_font, fmt=money_format)

# Combustible (fila 17)
apply_style(ws, 17, 2, 'Combustible', label_font)
apply_style(ws, 17, 3, '', data_font)
apply_style(ws, 17, 4, font=data_font, fmt=money_format)
ws.cell(row=17, column=4).value = "=COMBUSTIBLE!A2"  # FORMULA
apply_style(ws, 17, 5, '', data_font)
apply_style(ws, 17, 6, '', data_font)
apply_style(ws, 17, 7, font=data_font, fmt=money_format)

# Imprevistos (fila 18)
apply_style(ws, 18, 2, 'Imprevistos', label_font)
apply_style(ws, 18, 3, '', data_font)
apply_style(ws, 18, 4, font=data_font, fmt=money_format)
ws.cell(row=18, column=4).value = "='% IMPREVISTOS'!A2"  # FORMULA
apply_style(ws, 18, 5, '', data_font)
apply_style(ws, 18, 6, '', data_font)
apply_style(ws, 18, 7, font=data_font, fmt=money_format)

# Tecnicos asignados
apply_style(ws, 19, 2, 'Tecnico Asignado', label_font, light_fill)
apply_style(ws, 19, 3, '', data_font, light_fill)
apply_style(ws, 19, 4, font=data_font, fill=light_fill, fmt=money_format)

# Leer tecnicos del original
tecnicos = []
for orig_r in range(19, 26):
    name = rs.cell_value(orig_r, 1)
    if name and name.strip():
        val = rs.cell_value(orig_r, 2) if rs.cell_type(orig_r, 2) == 2 else 0
        tecnicos.append((name, val))

tec_start_row = 20
for i, (name, val) in enumerate(tecnicos):
    r = tec_start_row + i
    apply_style(ws, r, 2, name, data_font)
    apply_style(ws, r, 3, val, data_font, fmt=money_format)
    apply_style(ws, r, 4, '', data_font)

tec_end_row = tec_start_row + len(tecnicos) - 1

# Fila en blanco
blank_row = tec_end_row + 1
for _ in range(3):
    blank_row += 1

# Fila TOTAL
total_row = tec_end_row + 2
apply_style(ws, total_row, 2, 'Total', Font(name='Arial', size=11, bold=True), total_fill)
# C: Total viaticos no rendibles
apply_style(ws, total_row, 3, font=Font(name='Arial', size=11, bold=True), fill=total_fill, fmt=money_format)
ws.cell(row=total_row, column=3).value = f"=SUM(C{tec_start_row}:C{tec_end_row})"

# D: Total fondo a rendir
apply_style(ws, total_row, 4, font=Font(name='Arial', size=11, bold=True), fill=total_fill, fmt=money_format)
ws.cell(row=total_row, column=4).value = "=SUM(D14:D18)"

# E: Total gastos representacion
apply_style(ws, total_row, 5, font=Font(name='Arial', size=11, bold=True), fill=total_fill, fmt=money_format)
ws.cell(row=total_row, column=5).value = "=SUM(E14:E18)"

# F: Total reembolso
apply_style(ws, total_row, 6, font=Font(name='Arial', size=11, bold=True), fill=total_fill, fmt=money_format)
ws.cell(row=total_row, column=6).value = "=SUM(F14:F18)"

# G: TOTAL GENERAL
apply_style(ws, total_row, 7, font=Font(name='Arial', size=11, bold=True, color='FFFFFF'), fill=PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid'), fmt=money_format)
ws.cell(row=total_row, column=7).value = f"=C{total_row}+D{total_row}+E{total_row}+F{total_row}"

# Seccion vehiculo/logistica
log_row = total_row + 2
apply_style(ws, log_row, 2, 'Combustible', label_font, light_fill)
apply_style(ws, log_row, 5, 'Dias', label_font, light_fill)
apply_style(ws, log_row, 6, 1, data_font, fmt='0')

apply_style(ws, log_row + 1, 2, 'Vehiculo Empresa', label_font)
apply_style(ws, log_row + 1, 3, 'X', data_font, alignment=Alignment(horizontal='center'))
apply_style(ws, log_row + 1, 5, 'Noches', label_font)
apply_style(ws, log_row + 1, 6, 0, data_font, fmt='0')

apply_style(ws, log_row + 2, 2, 'Vehiculo arrendado', label_font)
apply_style(ws, log_row + 2, 3, '', data_font, alignment=Alignment(horizontal='center'))

apply_style(ws, log_row + 3, 2, 'Vehiculo propio', label_font)
apply_style(ws, log_row + 3, 3, '', data_font, alignment=Alignment(horizontal='center'))

# Nacional/Internacional
ni_row = log_row + 5
apply_style(ws, ni_row, 2, 'Nacional', label_font)
apply_style(ws, ni_row, 3, 'X', data_font, alignment=Alignment(horizontal='center'))
apply_style(ws, ni_row, 5, 'Fondo a rendir', label_font, total_fill)
apply_style(ws, ni_row, 6, 'Viatico no rendible', label_font, total_fill)

apply_style(ws, ni_row + 1, 2, 'Internacional', label_font)
apply_style(ws, ni_row + 1, 3, '', data_font, alignment=Alignment(horizontal='center'))
apply_style(ws, ni_row + 1, 5, 'rendir', Font(name='Arial', size=9, italic=True))
apply_style(ws, ni_row + 1, 6, 'Rendible', Font(name='Arial', size=9, italic=True))

# Resumen con formulas
apply_style(ws, ni_row + 2, 5, font=Font(name='Arial', size=11, bold=True), fill=total_fill, fmt=money_format)
ws.cell(row=ni_row + 2, column=5).value = f"=D{total_row}"  # Fondo a rendir
apply_style(ws, ni_row + 2, 6, font=Font(name='Arial', size=11, bold=True), fill=total_fill, fmt=money_format)
ws.cell(row=ni_row + 2, column=6).value = f"=C{total_row}"  # Viatico no rendible

# Fechas
fecha_row = ni_row + 4
import datetime
fecha_val = rs.cell_value(37, 2)
if rs.cell_type(37, 2) == 3:
    fecha_inicio = datetime.datetime(*xlrd.xldate_as_tuple(fecha_val, rb.datemode))
else:
    fecha_inicio = datetime.datetime(2026, 2, 21)

fecha_val2 = rs.cell_value(38, 2)
if rs.cell_type(38, 2) == 3:
    fecha_termino = datetime.datetime(*xlrd.xldate_as_tuple(fecha_val2, rb.datemode))
else:
    fecha_termino = fecha_inicio

apply_style(ws, fecha_row, 2, 'Fecha de Inicio', label_font, light_fill)
apply_style(ws, fecha_row, 3, fecha_inicio, data_font, fmt=date_format)
apply_style(ws, fecha_row + 1, 2, 'Fecha de termino', label_font, light_fill)
apply_style(ws, fecha_row + 1, 3, fecha_termino, data_font, fmt=date_format)

# Destinos
dest_row = fecha_row + 3
for i in range(1, 5):
    apply_style(ws, dest_row + i - 1, 2, f'Destino {i}', label_font, light_fill)
    orig_val = rs.cell_value(39 + i, 2) if rs.cell_type(39 + i, 2) == 1 else ''
    apply_style(ws, dest_row + i - 1, 3, orig_val, data_font)

# Jefe de area
jefe_row = dest_row + 5
apply_style(ws, jefe_row, 2, 'Jefe de Area', label_font, PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'))
ws.cell(row=jefe_row, column=2).font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
apply_style(ws, jefe_row, 3, rs.cell_value(45, 2) if rs.cell_type(45, 2) == 1 else '', data_font)


# =========================================================================
# HOJA 2: PEAJE
# =========================================================================
ws_p = wb.create_sheet('PEAJE')
ps = rb.sheet_by_index(1)

ws_p.column_dimensions['A'].width = 35
ws_p.column_dimensions['B'].width = 20
ws_p.column_dimensions['C'].width = 20
ws_p.column_dimensions['D'].width = 18
ws_p.column_dimensions['E'].width = 5
ws_p.column_dimensions['F'].width = 20
ws_p.column_dimensions['G'].width = 18

# Titulo y total
apply_style(ws_p, 1, 1, 'Peaje Ida y Vuelta', header_font, header_fill)
ws_p.cell(row=1, column=1).font = header_font_white
apply_style(ws_p, 2, 1, font=Font(name='Arial', size=12, bold=True), fill=total_fill, fmt=money_format)
# A2 = total de peaje seleccionado (se debe ingresar manualmente o seleccionar)
# Por defecto dejamos una referencia editable
ws_p.cell(row=2, column=1).value = 30000  # Valor por defecto (editable)

# Encabezados tabla
apply_style(ws_p, 5, 3, 'En Auto/Camioneta', header_font_white, header_fill, alignment=Alignment(horizontal='center'))
apply_style(ws_p, 5, 4, 'Ida y Vuelta', header_font_white, header_fill, alignment=Alignment(horizontal='center'))
apply_style(ws_p, 5, 6, 'En Camion', header_font_white, header_fill, alignment=Alignment(horizontal='center'))
apply_style(ws_p, 5, 7, 'Ida y Vuelta', header_font_white, header_fill, alignment=Alignment(horizontal='center'))

# Datos de peaje por destino (norte)
destinos_norte = []
for r in range(5, 12):
    ciudad = ps.cell_value(r, 1) if ps.cell_type(r, 1) == 1 else ''
    if ciudad:
        auto_ida = ps.cell_value(r, 2) if ps.cell_type(r, 2) == 2 else 0
        cam_ida = ps.cell_value(r, 5) if ps.cell_type(r, 5) == 2 else 0
        destinos_norte.append((ciudad, auto_ida, cam_ida))

excel_row = 6
apply_style(ws_p, excel_row, 1, 'Santiago', label_font, light_fill)

for ciudad, auto, cam in destinos_norte:
    apply_style(ws_p, excel_row, 2, ciudad, data_font)
    apply_style(ws_p, excel_row, 3, auto, data_font, fmt=money_format)
    # Ida y Vuelta = solo ida * 2 (FORMULA)
    ws_p.cell(row=excel_row, column=4).value = f"=C{excel_row}*2"
    apply_style(ws_p, excel_row, 4, font=data_font, fmt=money_format)
    apply_style(ws_p, excel_row, 6, cam, data_font, fmt=money_format)
    ws_p.cell(row=excel_row, column=7).value = f"=F{excel_row}*2"
    apply_style(ws_p, excel_row, 7, font=data_font, fmt=money_format)
    excel_row += 1

excel_row += 1

# Destinos sur
destinos_sur = []
for r in range(13, 22):
    ciudad = ps.cell_value(r, 1) if ps.cell_type(r, 1) == 1 else ''
    if ciudad:
        auto_ida = ps.cell_value(r, 2) if ps.cell_type(r, 2) == 2 else 0
        cam_ida = ps.cell_value(r, 5) if ps.cell_type(r, 5) == 2 else 0
        destinos_sur.append((ciudad, auto_ida, cam_ida))

apply_style(ws_p, excel_row, 1, 'Santiago', label_font, light_fill)
for ciudad, auto, cam in destinos_sur:
    apply_style(ws_p, excel_row, 2, ciudad, data_font)
    apply_style(ws_p, excel_row, 3, auto, data_font, fmt=money_format)
    ws_p.cell(row=excel_row, column=4).value = f"=C{excel_row}*2"
    apply_style(ws_p, excel_row, 4, font=data_font, fmt=money_format)
    apply_style(ws_p, excel_row, 6, cam, data_font, fmt=money_format)
    ws_p.cell(row=excel_row, column=7).value = f"=F{excel_row}*2"
    apply_style(ws_p, excel_row, 7, font=data_font, fmt=money_format)
    excel_row += 1

excel_row += 2

# Peajes individuales norte
apply_style(ws_p, excel_row, 1, '', data_font)
apply_style(ws_p, excel_row + 1, 2, 'Valor', label_font, light_fill)
apply_style(ws_p, excel_row + 1, 3, 'Valor', label_font, light_fill)
excel_row += 2

apply_style(ws_p, excel_row, 1, 'Peajes hacia el norte', label_font, header_fill)
ws_p.cell(row=excel_row, column=1).font = header_font_white
apply_style(ws_p, excel_row, 2, 'Auto/Camioneta', header_font_white, header_fill)
apply_style(ws_p, excel_row, 3, 'Camion', header_font_white, header_fill)
excel_row += 1

peajes_norte = []
for r in range(26, 32):
    nombre = ps.cell_value(r, 0) if ps.cell_type(r, 0) == 1 else ''
    if nombre:
        auto = ps.cell_value(r, 1) if ps.cell_type(r, 1) == 2 else 0
        cam = ps.cell_value(r, 2) if ps.cell_type(r, 2) == 2 else 0
        peajes_norte.append((nombre, auto, cam))

for nombre, auto, cam in peajes_norte:
    apply_style(ws_p, excel_row, 1, nombre, data_font)
    apply_style(ws_p, excel_row, 2, auto, data_font, fmt=money_format)
    apply_style(ws_p, excel_row, 3, cam, data_font, fmt=money_format)
    excel_row += 1

excel_row += 1

# Peajes individuales sur
apply_style(ws_p, excel_row, 1, 'Peajes hacia el Sur', label_font, header_fill)
ws_p.cell(row=excel_row, column=1).font = header_font_white
apply_style(ws_p, excel_row, 2, 'Auto/Camioneta', header_font_white, header_fill)
apply_style(ws_p, excel_row, 3, 'Camion', header_font_white, header_fill)
excel_row += 1

peajes_sur = []
for r in range(34, 46):
    nombre = ps.cell_value(r, 0) if ps.cell_type(r, 0) == 1 else ''
    if nombre:
        auto = ps.cell_value(r, 1) if ps.cell_type(r, 1) == 2 else 0
        cam = ps.cell_value(r, 2) if ps.cell_type(r, 2) == 2 else 0
        peajes_sur.append((nombre, auto, cam))

for nombre, auto, cam in peajes_sur:
    apply_style(ws_p, excel_row, 1, nombre, data_font)
    apply_style(ws_p, excel_row, 2, auto, data_font, fmt=money_format)
    apply_style(ws_p, excel_row, 3, cam, data_font, fmt=money_format)
    excel_row += 1


# =========================================================================
# HOJA 3: COMBUSTIBLE
# =========================================================================
ws_c = wb.create_sheet('COMBUSTIBLE')
cs = rb.sheet_by_index(2)

ws_c.column_dimensions['A'].width = 35
ws_c.column_dimensions['B'].width = 20
ws_c.column_dimensions['C'].width = 20

# Titulo y total con FORMULA
apply_style(ws_c, 1, 1, 'Combustible Ida y Vuelta', header_font, header_fill)
ws_c.cell(row=1, column=1).font = header_font_white
apply_style(ws_c, 2, 1, font=Font(name='Arial', size=12, bold=True), fill=total_fill, fmt=money_format)
ws_c.cell(row=2, column=1).value = "=B18"  # FORMULA: Total = costo auto

# Encabezados
apply_style(ws_c, 7, 2, 'Auto/Camioneta', header_font_white, header_fill, alignment=Alignment(horizontal='center'))
apply_style(ws_c, 7, 3, 'Camion', header_font_white, header_fill, alignment=Alignment(horizontal='center'))

# Datos
apply_style(ws_c, 8, 1, 'Valor Combustible por litro ($)', label_font, light_fill)
apply_style(ws_c, 8, 2, 1450, data_font, fmt=money_format)
apply_style(ws_c, 8, 3, 1100, data_font, fmt=money_format)

apply_style(ws_c, 9, 1, 'Consumo (Km/Lt)', label_font, light_fill)
apply_style(ws_c, 9, 2, 10, data_font, fmt='0')
apply_style(ws_c, 9, 3, 6, data_font, fmt='0')

apply_style(ws_c, 10, 1, 'Kilometros de ida', label_font)
apply_style(ws_c, 10, 2, 800, data_font, fmt='#,##0')
apply_style(ws_c, 10, 3, 800, data_font, fmt='#,##0')

apply_style(ws_c, 11, 1, 'Kilometros de vuelta', label_font)
apply_style(ws_c, 11, 2, 800, data_font, fmt='#,##0')
apply_style(ws_c, 11, 3, 800, data_font, fmt='#,##0')

# KM totales (FORMULA)
apply_style(ws_c, 12, 1, 'Kilometros totales', label_font, total_fill)
apply_style(ws_c, 12, 2, font=Font(name='Arial', size=10, bold=True), fill=total_fill, fmt='#,##0')
ws_c.cell(row=12, column=2).value = "=B10+B11"  # FORMULA
apply_style(ws_c, 12, 3, font=Font(name='Arial', size=10, bold=True), fill=total_fill, fmt='#,##0')
ws_c.cell(row=12, column=3).value = "=C10+C11"  # FORMULA

# Consumo total lts (FORMULA)
apply_style(ws_c, 13, 1, 'Consumo combustible total (lts.)', label_font)
apply_style(ws_c, 13, 2, font=data_font, fmt='#,##0.0')
ws_c.cell(row=13, column=2).value = "=B12/B9"  # FORMULA
apply_style(ws_c, 13, 3, font=data_font, fmt='#,##0.0')
ws_c.cell(row=13, column=3).value = "=C12/C9"  # FORMULA

# Encabezados resultado
apply_style(ws_c, 17, 2, 'Auto/Camioneta', header_font_white, header_fill, alignment=Alignment(horizontal='center'))
apply_style(ws_c, 17, 3, 'Camion', header_font_white, header_fill, alignment=Alignment(horizontal='center'))

# Costo total (FORMULA)
apply_style(ws_c, 18, 1, 'Costo combustible total ($)', label_font, total_fill)
apply_style(ws_c, 18, 2, font=Font(name='Arial', size=11, bold=True), fill=total_fill, fmt=money_format)
ws_c.cell(row=18, column=2).value = "=B13*B8"  # FORMULA
apply_style(ws_c, 18, 3, font=Font(name='Arial', size=11, bold=True), fill=total_fill, fmt=money_format)
ws_c.cell(row=18, column=3).value = "=C13*C8"  # FORMULA


# =========================================================================
# HOJA 4: ALOJAMIENTO EN CAMINO
# =========================================================================
ws_a = wb.create_sheet('ALOJAMIENTO EN CAMINO')
als = rb.sheet_by_index(3)

ws_a.column_dimensions['A'].width = 5
ws_a.column_dimensions['B'].width = 35
ws_a.column_dimensions['C'].width = 15
ws_a.column_dimensions['D'].width = 15
ws_a.column_dimensions['E'].width = 15

# Titulo y total
apply_style(ws_a, 1, 1, 'Total alojamiento', header_font, header_fill)
ws_a.cell(row=1, column=1).font = header_font_white
apply_style(ws_a, 2, 1, font=Font(name='Arial', size=12, bold=True), fill=total_fill, fmt=money_format)
ws_a.cell(row=2, column=1).value = "=D15"  # FORMULA: total promedio

# Encabezados
apply_style(ws_a, 6, 4, 'Valores', header_font_white, header_fill, alignment=Alignment(horizontal='center'))
apply_style(ws_a, 7, 3, 'bajo', label_font, light_fill, alignment=Alignment(horizontal='center'))
apply_style(ws_a, 7, 4, 'promedio', label_font, light_fill, alignment=Alignment(horizontal='center'))
apply_style(ws_a, 7, 5, 'alto', label_font, light_fill, alignment=Alignment(horizontal='center'))

apply_style(ws_a, 8, 2, 'Dias ida', label_font)
apply_style(ws_a, 8, 4, 3, data_font, fmt='0')

apply_style(ws_a, 9, 2, 'Dias vuelta', label_font)
apply_style(ws_a, 9, 4, 0, data_font, fmt='0')

apply_style(ws_a, 10, 2, 'Habitacion doble con bano', label_font)
apply_style(ws_a, 10, 3, 20000, data_font, fmt=money_format)
apply_style(ws_a, 10, 4, 70000, data_font, fmt=money_format)
apply_style(ws_a, 10, 5, 60000, data_font, fmt=money_format)

apply_style(ws_a, 11, 2, 'Habitacion single con bano', label_font)
apply_style(ws_a, 11, 3, '', data_font)
apply_style(ws_a, 11, 4, '', data_font)
apply_style(ws_a, 11, 5, '', data_font)

# Total alojamiento (FORMULAS)
apply_style(ws_a, 15, 2, 'Total alojamiento en camino', label_font, total_fill)
apply_style(ws_a, 15, 3, font=Font(name='Arial', size=10, bold=True), fill=total_fill, fmt=money_format)
ws_a.cell(row=15, column=3).value = "=(D8+D9)*C10"  # FORMULA bajo
apply_style(ws_a, 15, 4, font=Font(name='Arial', size=10, bold=True), fill=total_fill, fmt=money_format)
ws_a.cell(row=15, column=4).value = "=(D8+D9)*D10"  # FORMULA promedio
apply_style(ws_a, 15, 5, font=Font(name='Arial', size=10, bold=True), fill=total_fill, fmt=money_format)
ws_a.cell(row=15, column=5).value = "=(D8+D9)*E10"  # FORMULA alto


# =========================================================================
# HOJA 5: % IMPREVISTOS
# =========================================================================
ws_i = wb.create_sheet('% IMPREVISTOS')

ws_i.column_dimensions['A'].width = 10
ws_i.column_dimensions['B'].width = 35
ws_i.column_dimensions['C'].width = 15
ws_i.column_dimensions['D'].width = 15
ws_i.column_dimensions['E'].width = 20

# Titulo y total
apply_style(ws_i, 1, 1, '20% Imprevistos', header_font, header_fill)
ws_i.cell(row=1, column=1).font = header_font_white
apply_style(ws_i, 2, 1, font=Font(name='Arial', size=12, bold=True), fill=total_fill, fmt=money_format)
ws_i.cell(row=2, column=1).value = "=E4"  # FORMULA

# Configuracion
apply_style(ws_i, 4, 1, 0.2, data_font, fmt=pct_format)  # 20%
apply_style(ws_i, 4, 2, '20% PARA GASTOS IMPREVISTOS', label_font)
apply_style(ws_i, 4, 5, font=Font(name='Arial', size=11, bold=True), fill=total_fill, fmt=money_format)
ws_i.cell(row=4, column=5).value = "=A4*(PEAJE!A2+COMBUSTIBLE!A2+'ALOJAMIENTO EN CAMINO'!A2)"  # FORMULA


# =========================================================================
# HOJA 6: KILOMETRAJE
# =========================================================================
ws_k = wb.create_sheet('KILOMETRAJE')
ks = rb.sheet_by_index(5)

ws_k.column_dimensions['A'].width = 5
ws_k.column_dimensions['B'].width = 25
ws_k.column_dimensions['C'].width = 15

apply_style(ws_k, 1, 2, 'Distancias para ingresar en Combustible', header_font, header_fill)
ws_k.cell(row=1, column=2).font = header_font_white
apply_style(ws_k, 1, 3, 'Km desde Stgo', header_font_white, header_fill, alignment=Alignment(horizontal='center'))

for r in range(4, 31):
    ciudad = ks.cell_value(r, 1) if ks.cell_type(r, 1) == 1 else ''
    km = ks.cell_value(r, 2) if ks.cell_type(r, 2) == 2 else ''
    if ciudad:
        row_excel = r - 2
        apply_style(ws_k, row_excel, 2, ciudad, data_font, light_fill if (r % 2 == 0) else None)
        if km:
            apply_style(ws_k, row_excel, 3, km, data_font, fmt='#,##0')


# =========================================================================
# GUARDAR
# =========================================================================
wb.save(OUTPUT_FILE)
print(f"[OK] Archivo corregido guardado como: {OUTPUT_FILE}")
print()
print("FORMULAS VINCULADAS:")
print("=" * 60)
print()
print("HOJA SOLICITUD:")
print("  D14 (Peajes)        = PEAJE!A2")
print("  D16 (Hotel)         = 'ALOJAMIENTO EN CAMINO'!A2")
print("  D17 (Combustible)   = COMBUSTIBLE!A2")
print("  D18 (Imprevistos)   = '% IMPREVISTOS'!A2")
print(f"  C{total_row} (Total viatico) = SUM(C{tec_start_row}:C{tec_end_row})")
print(f"  D{total_row} (Total rendir)  = SUM(D14:D18)")
print(f"  E{total_row} (Total repres.) = SUM(E14:E18)")
print(f"  F{total_row} (Total reemb.)  = SUM(F14:F18)")
print(f"  G{total_row} (TOTAL GENERAL) = C{total_row}+D{total_row}+E{total_row}+F{total_row}")
print()
print("HOJA COMBUSTIBLE:")
print("  A2  = B18     (total combustible)")
print("  B12 = B10+B11 (km totales Auto)")
print("  C12 = C10+C11 (km totales Camion)")
print("  B13 = B12/B9  (consumo lts Auto)")
print("  C13 = C12/C9  (consumo lts Camion)")
print("  B18 = B13*B8  (costo total Auto)")
print("  C18 = C13*C8  (costo total Camion)")
print()
print("HOJA ALOJAMIENTO EN CAMINO:")
print("  A2  = D15          (total alojamiento promedio)")
print("  C15 = (D8+D9)*C10  (total bajo)")
print("  D15 = (D8+D9)*D10  (total promedio)")
print("  E15 = (D8+D9)*E10  (total alto)")
print()
print("HOJA % IMPREVISTOS:")
print("  A2  = E4")
print("  E4  = A4 * (PEAJE!A2 + COMBUSTIBLE!A2 + 'ALOJAMIENTO EN CAMINO'!A2)")
print()
print("HOJA PEAJE:")
print("  Columnas D y G: Ida y Vuelta = Solo ida * 2 (formulas)")
