"""
Gestión de datos para la app de Viáticos.
Maneja la lectura/escritura de archivos JSON de configuración y solicitudes.
"""

import json
import os
import sys
import copy
from datetime import date, datetime
import io

def get_base_dir():
    # Detecta si se está ejecutando como archivo .exe compilado por PyInstaller
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.dirname(__file__))

BASE_DIR = get_base_dir()
CONFIG_DIR = os.path.join(BASE_DIR, 'config')
DEFAULT_DATA_PATH = os.path.join(CONFIG_DIR, 'default_data.json')
SOLICITUD_PATH = os.path.join(CONFIG_DIR, 'solicitud_actual.json')
HISTORIAL_DIR = os.path.join(CONFIG_DIR, 'historial')


def ensure_dirs():
    """Asegura que existan los directorios necesarios."""
    os.makedirs(CONFIG_DIR, exist_ok=True)
    os.makedirs(HISTORIAL_DIR, exist_ok=True)


def load_default_data():
    """Carga los datos de referencia por defecto."""
    with open(DEFAULT_DATA_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)


def save_default_data(data):
    """Guarda los datos de referencia actualizados."""
    with open(DEFAULT_DATA_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def get_empty_solicitud():
    """Retorna una solicitud vacía con estructura completa."""
    return {
        "fecha_solicitud": date.today().isoformat(),
        "cliente": "",
        "motivo": "",
        "acn": "",
        "orden_compra": "",
        "solicitado_por": "",
        "responsable": "",
        "jefe_area": "Cristián Sánchez Rojas",
        "tipo_viaje": "Nacional",
        "tipo_vehiculo": "Vehiculo Empresa",
        "tipo_auto_camion": "Auto/Camioneta",
        "ida_vuelta": True,
        "dias": 1,
        "noches": 0,
        "fecha_inicio": date.today().isoformat(),
        "fecha_termino": date.today().isoformat(),
        "destinos": ["", "", "", ""],
        "tecnicos": [],
        "tipo_habitacion": "habitacion_doble",
        "rango_precio_alojamiento": "promedio",
    }


class DateEncoder(json.JSONEncoder):
    """Encoder para serializar objetos date/datetime."""
    def default(self, obj):
        if isinstance(obj, (date, datetime)):
            return obj.isoformat()
        return super().default(obj)


def load_solicitud():
    """Carga la solicitud actual. Si no existe, crea una vacía."""
    ensure_dirs()
    if os.path.exists(SOLICITUD_PATH):
        with open(SOLICITUD_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
            empty = get_empty_solicitud()
            empty.update(data)
            return empty
    return get_empty_solicitud()


def save_solicitud(data):
    """Guarda la solicitud actual."""
    ensure_dirs()
    with open(SOLICITUD_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False, cls=DateEncoder)


def save_to_historial(solicitud, ref_data, calculos):
    """Guarda una copia de la solicitud en el historial."""
    ensure_dirs()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    cliente = solicitud.get('cliente', 'sin_cliente').replace(' ', '_')[:20]
    filename = f"solicitud_{timestamp}_{cliente}.json"
    filepath = os.path.join(HISTORIAL_DIR, filename)
    
    record = {
        "solicitud": solicitud,
        "ref_data_snapshot": ref_data,
        "calculos": calculos,
        "timestamp": datetime.now().isoformat()
    }
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(record, f, indent=2, ensure_ascii=False, cls=DateEncoder)
    
    return filepath


def list_historial():
    """Lista las solicitudes guardadas en el historial."""
    ensure_dirs()
    files = []
    for f in sorted(os.listdir(HISTORIAL_DIR), reverse=True):
        if f.endswith('.json'):
            filepath = os.path.join(HISTORIAL_DIR, f)
            try:
                with open(filepath, 'r', encoding='utf-8') as fh:
                    data = json.load(fh)
                    files.append({
                        "filename": f,
                        "filepath": filepath,
                        "timestamp": data.get("timestamp", ""),
                        "cliente": data.get("solicitud", {}).get("cliente", ""),
                        "motivo": data.get("solicitud", {}).get("motivo", ""),
                        "destino": data.get("solicitud", {}).get("destinos", [""])[0],
                        "total": data.get("calculos", {}).get("total_general", 0)
                    })
            except Exception:
                pass
    return files


def load_historial_entry(filepath):
    """Carga una entrada del historial completa."""
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_ciudades_destino(ref_data):
    """Retorna lista de todas las ciudades disponibles como destino."""
    ciudades = set()
    for item in ref_data.get("peajes_por_destino", {}).get("norte", []):
        ciudades.add(item["ciudad"])
    for item in ref_data.get("peajes_por_destino", {}).get("sur", []):
        ciudades.add(item["ciudad"])
    for item in ref_data.get("kilometraje", []):
        if item["ciudad"] != "Santiago":
            ciudades.add(item["ciudad"])
    return sorted(list(ciudades))


def get_km_for_ciudad(ref_data, ciudad):
    """Retorna los km desde Santiago para una ciudad dada."""
    for item in ref_data.get("kilometraje", []):
        if item["ciudad"].lower() == ciudad.lower():
            return item["km"]
    return 0


def generar_excel(solicitud, calculos, ref_data):
    """
    Genera un archivo Excel con el mismo formato que la pestaña 
    'Solicitud' del documento original. Retorna bytes del archivo.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solicitud"
    
    # Estilos
    header_font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    title_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    label_font = Font(name='Arial', bold=True, size=10)
    value_font = Font(name='Arial', size=10)
    money_font = Font(name='Arial', bold=True, size=11, color='1F4E79')
    
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    alt_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    total_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18
    
    # ── Encabezado ──
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = "SOLICITUD DE VIATICOS Y FONDOS"
    cell.font = header_font
    cell.alignment = center
    
    ws.merge_cells('A2:E2')
    cell = ws['A2']
    cell.value = "Metrica Spa"
    cell.font = Font(name='Arial', bold=True, size=11, color='4472C4')
    cell.alignment = center
    
    # ── Datos Generales ──
    row = 4
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "DATOS GENERALES"
    cell.font = title_font
    cell.fill = header_fill
    cell.alignment = center
    
    campos_generales = [
        ("Cliente:", solicitud.get("cliente", "")),
        ("ACN:", solicitud.get("acn", "")),
        ("Motivo:", solicitud.get("motivo", "")),
        ("Solicitado por:", solicitud.get("solicitado_por", "")),
        ("Responsable:", solicitud.get("responsable", "")),
        ("Jefe de Area:", solicitud.get("jefe_area", "")),
        ("Orden de Compra:", solicitud.get("orden_compra", "")),
    ]
    
    for label, value in campos_generales:
        row += 1
        ws[f'A{row}'].value = label
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].border = thin_border
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'].value = value
        ws[f'B{row}'].font = value_font
        ws[f'B{row}'].border = thin_border
    
    # ── Datos del Viaje ──
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "DATOS DEL VIAJE"
    cell.font = title_font
    cell.fill = header_fill
    cell.alignment = center
    
    destinos = solicitud.get("destinos", ["", "", "", ""])
    dest_str = ", ".join([d for d in destinos if d])
    
    campos_viaje = [
        ("Tipo de viaje:", solicitud.get("tipo_viaje", "Nacional")),
        ("Vehiculo:", solicitud.get("tipo_vehiculo", "")),
        ("Tipo vehiculo:", solicitud.get("tipo_auto_camion", "")),
        ("Ida y vuelta:", "Si" if solicitud.get("ida_vuelta", True) else "Solo ida"),
        ("Destino(s):", dest_str),
        ("Fecha inicio:", solicitud.get("fecha_inicio", "")),
        ("Fecha termino:", solicitud.get("fecha_termino", "")),
        ("Dias:", str(solicitud.get("dias", 0))),
        ("Noches:", str(solicitud.get("noches", 0))),
        ("Tipo habitacion:", "Doble" if solicitud.get("tipo_habitacion") == "habitacion_doble" else "Single"),
    ]
    
    for label, value in campos_viaje:
        row += 1
        ws[f'A{row}'].value = label
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].border = thin_border
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'].value = value
        ws[f'B{row}'].font = value_font
        ws[f'B{row}'].border = thin_border
    
    # ── Técnicos ──
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "TECNICOS ASIGNADOS"
    cell.font = title_font
    cell.fill = header_fill
    cell.alignment = center
    
    row += 1
    for col, header in [('A', 'Tecnico'), ('B', 'Viatico No Rendible')]:
        ws[f'{col}{row}'].value = header
        ws[f'{col}{row}'].font = label_font
        ws[f'{col}{row}'].fill = alt_fill
        ws[f'{col}{row}'].border = thin_border
        ws[f'{col}{row}'].alignment = center
    
    tecnicos = solicitud.get("tecnicos", [])
    for t in tecnicos:
        row += 1
        ws[f'A{row}'].value = t["nombre"]
        ws[f'A{row}'].font = value_font
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].value = t["monto"]
        ws[f'B{row}'].font = value_font
        ws[f'B{row}'].border = thin_border
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'B{row}'].alignment = right_align
    
    # ── Desglose de Costos ──
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "DESGLOSE DE COSTOS"
    cell.font = title_font
    cell.fill = header_fill
    cell.alignment = center
    
    row += 1
    headers_cost = ['Concepto', 'Fondo a Rendir', 'Viatico No Rendible']
    for i, h in enumerate(headers_cost):
        col = chr(65 + i)
        ws[f'{col}{row}'].value = h
        ws[f'{col}{row}'].font = label_font
        ws[f'{col}{row}'].fill = alt_fill
        ws[f'{col}{row}'].border = thin_border
        ws[f'{col}{row}'].alignment = center
    
    desglose_rows = [
        ("Peajes", calculos["peaje"]["peaje_ida_vuelta"], 0),
        ("Alojamiento", calculos["alojamiento"]["total"], 0),
        ("Combustible", calculos["combustible"]["costo_total"], 0),
        ("Imprevistos", calculos["imprevistos"]["total_imprevistos"], 0),
        ("Viaticos Tecnicos", 0, calculos["total_viatico_no_rendible"]),
    ]
    
    for concepto, fondo, viatico in desglose_rows:
        row += 1
        ws[f'A{row}'].value = concepto
        ws[f'A{row}'].font = value_font
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].value = fondo
        ws[f'B{row}'].font = value_font
        ws[f'B{row}'].border = thin_border
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'B{row}'].alignment = right_align
        ws[f'C{row}'].value = viatico
        ws[f'C{row}'].font = value_font
        ws[f'C{row}'].border = thin_border
        ws[f'C{row}'].number_format = '#,##0'
        ws[f'C{row}'].alignment = right_align
    
    # Total
    row += 1
    ws[f'A{row}'].value = "TOTAL GENERAL"
    ws[f'A{row}'].font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    ws[f'A{row}'].fill = total_fill
    ws[f'A{row}'].border = thin_border
    ws.merge_cells(f'B{row}:C{row}')
    ws[f'B{row}'].value = calculos["total_general"]
    ws[f'B{row}'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    ws[f'B{row}'].fill = total_fill
    ws[f'B{row}'].border = thin_border
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'B{row}'].alignment = center
    
    # Guardar a BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generar_pdf(solicitud, calculos, ref_data):
    """
    Genera un PDF profesional con el resumen de la solicitud usando xhtml2pdf.
    Retorna bytes del archivo PDF generado.
    """
    import io
    from xhtml2pdf import pisa
    
    tecnicos = solicitud.get("tecnicos", [])
    destinos = solicitud.get("destinos", ["", "", "", ""])
    dest_str = ", ".join([d for d in destinos if d]) or "—"
    
    tec_rows = ""
    for t in tecnicos:
        tec_rows += f"<tr><td>{t['nombre']}</td><td class='money'>${t['monto']:,.0f}</td></tr>"
    
    desglose_rows = f"""
    <tr><td>Peajes</td><td class='money'>${calculos['peaje']['peaje_ida_vuelta']:,.0f}</td></tr>
    <tr><td>Alojamiento</td><td class='money'>${calculos['alojamiento']['total']:,.0f}</td></tr>
    <tr><td>Combustible</td><td class='money'>${calculos['combustible']['costo_total']:,.0f}</td></tr>
    <tr><td>Imprevistos ({calculos['imprevistos']['porcentaje']*100:.0f}%)</td><td class='money'>${calculos['imprevistos']['total_imprevistos']:,.0f}</td></tr>
    <tr><td>Viaticos Tecnicos</td><td class='money'>${calculos['total_viatico_no_rendible']:,.0f}</td></tr>
    """
    
    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
    @page {{
        size: letter;
        margin: 1cm 1.5cm;
    }}
    @font-face {{
        font-family: 'Helvetica';
    }}
    body {{ font-family: Helvetica, sans-serif; color: #1a1a2e; font-size: 10px; padding: 0; margin: 0; }}
    .header {{ text-align: center; margin-bottom: 10px; border-bottom: 2px solid #1F4E79; padding-bottom: 10px; }}
    .header h1 {{ font-size: 16px; color: #1F4E79; font-weight: bold; margin-bottom: 2px; }}
    .header p {{ color: #666; font-size: 10px; }}
    .section-title {{ background-color: #1F4E79; color: white; padding: 3px 6px; font-size: 11px; font-weight: bold; margin-top: 10px; margin-bottom: 6px; }}
    table {{ width: 100%; border-collapse: collapse; margin-bottom: 6px; }}
    th, td {{ padding: 3px 6px; border: 1px solid #ddd; font-size: 10px; vertical-align: middle; }}
    th {{ background-color: #E8EEF4; font-weight: bold; text-align: left; }}
    td.money {{ text-align: right; font-weight: bold; }}
    .total-row td {{ background-color: #FFD700; font-size: 11px; font-weight: bold; color: #1F4E79; }}
    .kpi-table {{ width: 100%; border-collapse: separate; border-spacing: 5px; margin-top: 5px; margin-bottom: 5px; }}
    .kpi-table td {{ border: 1px solid #D0D8E8; background-color: #F0F4FA; text-align: center; padding: 5px; border-radius: 3px; }}
    .kpi-label {{ font-size: 9px; color: #666; }}
    .kpi-value {{ font-size: 12px; font-weight: bold; color: #1F4E79; margin-top: 2px; }}
    .total-box {{ background-color: #1F4E79; color: white; padding: 8px; text-align: center; margin-top: 10px; }}
    .total-box-label {{ font-size: 11px; }}
    .total-box-value {{ font-size: 18px; font-weight: bold; margin-top: 2px; }}
    .footer {{ text-align: center; color: #999; font-size: 8px; margin-top: 15px; border-top: 1px solid #eee; padding-top: 5px; }}
    
    /* Layout table for info grids since CSS Grid/Flex is not fully supported by xhtml2pdf */
    .layout-table {{ width: 100%; border: none; padding: 0; margin: 0; }}
    .layout-table td {{ border: none; padding: 2px 4px; font-size: 10px; }}
    .lbl {{ color: #666; font-weight: normal; }}
    .val {{ font-weight: bold; }}
</style></head><body>
    <div class="header">
        <h1>SOLICITUD DE VIATICOS Y FONDOS</h1>
        <p>Metrica Spa | Fecha: {solicitud.get('fecha_solicitud', '')}</p>
    </div>
    
    <div class="section-title">Datos Generales</div>
    <table class="layout-table">
        <tr>
            <td width="50%"><span class="lbl">Cliente:</span> <span class="val">{solicitud.get('cliente', '')}</span></td>
            <td width="50%"><span class="lbl">Solicitado por:</span> <span class="val">{solicitud.get('solicitado_por', '')}</span></td>
        </tr>
        <tr>
            <td><span class="lbl">ACN:</span> <span class="val">{solicitud.get('acn', '')}</span></td>
            <td><span class="lbl">Responsable:</span> <span class="val">{solicitud.get('responsable', '')}</span></td>
        </tr>
        <tr>
            <td><span class="lbl">Motivo:</span> <span class="val">{solicitud.get('motivo', '')}</span></td>
            <td><span class="lbl">Jefe de Area:</span> <span class="val">{solicitud.get('jefe_area', '')}</span></td>
        </tr>
    </table>
    
    <div class="section-title">Datos del Viaje</div>
    <table class="layout-table">
        <tr>
            <td width="50%"><span class="lbl">Destino(s):</span> <span class="val">{dest_str}</span></td>
            <td width="50%"><span class="lbl">Fecha inicio:</span> <span class="val">{solicitud.get('fecha_inicio', '')}</span></td>
        </tr>
        <tr>
            <td><span class="lbl">Tipo de viaje:</span> <span class="val">{solicitud.get('tipo_viaje', '')}</span></td>
            <td><span class="lbl">Fecha termino:</span> <span class="val">{solicitud.get('fecha_termino', '')}</span></td>
        </tr>
        <tr>
            <td><span class="lbl">Vehiculo:</span> <span class="val">{solicitud.get('tipo_vehiculo', '')}</span></td>
            <td><span class="lbl">Dias:</span> <span class="val">{solicitud.get('dias', 0)}</span></td>
        </tr>
        <tr>
            <td><span class="lbl">Ida y vuelta:</span> <span class="val">{'Si' if solicitud.get('ida_vuelta', True) else 'Solo ida'}</span></td>
            <td><span class="lbl">Noches:</span> <span class="val">{solicitud.get('noches', 0)}</span></td>
        </tr>
    </table>
    
    <div class="section-title">Tecnicos Asignados</div>
    <table>
        <tr><th>Tecnico</th><th style="text-align:right">Viatico No Rendible</th></tr>
        {tec_rows}
        <tr class="total-row"><td>Total Viaticos</td><td class="money">${calculos['total_viatico_no_rendible']:,.0f}</td></tr>
    </table>
    
    <table class="kpi-table">
        <tr>
            <td>
                <div class="kpi-label">Fondo a Rendir</div>
                <div class="kpi-value">${calculos['total_fondo_rendir']:,.0f}</div>
            </td>
            <td>
                <div class="kpi-label">Viatico No Rendible</div>
                <div class="kpi-value">${calculos['total_viatico_no_rendible']:,.0f}</div>
            </td>
            <td>
                <div class="kpi-label">Km Totales</div>
                <div class="kpi-value">{calculos['combustible']['km_totales']:,}</div>
            </td>
        </tr>
    </table>
    
    <div class="section-title">Desglose de Costos</div>
    <table>
        <tr><th>Concepto</th><th style="text-align:right">Monto</th></tr>
        {desglose_rows}
        <tr class="total-row"><td>TOTAL GENERAL</td><td class="money">${calculos['total_general']:,.0f}</td></tr>
    </table>
    
    <div class="total-box">
        <div class="total-box-label">Total General Solicitud</div>
        <div class="total-box-value">${calculos['total_general']:,.0f}</div>
    </div>
    
    <div class="footer">
        Documento generado automaticamente | Metrica Spa | {datetime.now().strftime('%d/%m/%Y %H:%M')}
    </div>
</body></html>"""
    
    output = io.BytesIO()
    # Generar el PDF
    pisa_status = pisa.CreatePDF(
        io.StringIO(html),
        dest=output
    )
    
    if pisa_status.err:
        raise Exception("Error al generar el documento PDF")
        
    return output.getvalue()
